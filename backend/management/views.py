import csv
import io
from openpyxl import load_workbook, Workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Count, Q
from django.utils import timezone
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime
from utils.auth import JWTCookieAuthentication, IsSuperUser
from users.models import Profile
from users.serializers import UserSerializer, ProfileSerializer, UserInverseSerializer
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from .serializers import LanguageSerializer, FrameworkSerializer, SettingSerializer
import io
from .models import Setting
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
from django.db.models import (
    Count,
    Q,
    F,
    Sum,
    When,
    Case,
    Value,
    FloatField,
    Avg,
)
from django.db.models.functions import TruncMonth
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import timedelta
import math
import environ
import cloudinary
import cloudinary.utils
from pathlib import Path
from asgiref.sync import async_to_sync
from utils.notif import notify_user, notify_users_bulk
from courses.models import Course, UserCourseProgress
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch


env = environ.Env()
BASE_DIR = Path(__file__).resolve().parent
environ.Env.read_env(str(BASE_DIR / ".env"))

cloudinary.config(
    cloud_name=env("CLOUD_NAME"),
    api_key=env("CLOUD_API_KEY"),
    api_secret=env("CLOUD_API_SECRET"),
    secure=True,
)

User = get_user_model()


class SimpleDataFrame:
    def __init__(self, rows, columns):
        """
        rows: a list of dictionaries, one per row
        columns: a list of column names
        """
        self._rows = rows
        self.columns = columns

    def iterrows(self):
        """
        Mimics pandas iterrows(): yields (index, row_dict)
        """
        for index, row in enumerate(self._rows):
            yield index, row


class StudentsView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAdminUser]

    def get(self, request):
        try:
            # Query params
            search = request.query_params.get("search", "").strip()
            grade = request.query_params.get("grade", "").strip()
            section = request.query_params.get("section", "").strip()
            field = request.query_params.get("field", "").strip()
            account_status = request.query_params.get("accountStatus", "").strip()
            sort_by = request.query_params.get("sort_by", "-user__date_joined")
            page = int(request.query_params.get("page", 1))
            page_size = int(request.query_params.get("page_size", 10))

            # Base queryset for all students
            base_profiles = Profile.objects.select_related("user").filter(
                user__role="user"
            )

            # Filter options
            all_grades = (
                base_profiles.exclude(grade__isnull=True)
                .values_list("grade", flat=True)
                .distinct()
                .order_by("grade")
            )
            all_sections = (
                base_profiles.exclude(section__isnull=True)
                .exclude(section="")
                .values_list("section", flat=True)
                .distinct()
                .order_by("section")
            )
            all_fields = (
                base_profiles.exclude(field__isnull=True)
                .exclude(field="")
                .values_list("field", flat=True)
                .distinct()
                .order_by("field")
            )

            filter_options = {
                "grades": list(all_grades),
                "sections": list(all_sections),
                "fields": list(all_fields),
            }

            # Annotate progress points
            profiles = base_profiles.annotate(
                total_courses=Count("user__course_progress__course", distinct=True),
                started_courses=Count(
                    "user__course_progress",
                    filter=Q(user__course_progress__status="started"),
                ),
                finished_courses=Count(
                    "user__course_progress",
                    filter=Q(user__course_progress__status="finished"),
                ),
            )

            # Progress percentage
            profiles = profiles.annotate(
                progress_percentage=Case(
                    When(total_courses=0, then=Value(0.0)),
                    default=100.0 * F("finished_courses") / F("total_courses"),
                    output_field=FloatField(),
                )
            )

            # Filters
            if search:
                search_filter = (
                    Q(user__full_name__icontains=search)
                    | Q(user__email__icontains=search)
                    | Q(grade__icontains=search)
                    | Q(section__icontains=search)
                    | Q(field__icontains=search)
                )
                profiles = profiles.filter(search_filter)

            if grade:
                try:
                    profiles = profiles.filter(grade=int(grade))
                except ValueError:
                    pass
            if section:
                profiles = profiles.filter(section__iexact=section)
            if field:
                profiles = profiles.filter(field__iexact=field)
            if account_status:
                if account_status.lower() == "active":
                    profiles = profiles.filter(user__is_active=True)
                elif account_status.lower() == "inactive":
                    profiles = profiles.filter(user__is_active=False)

            # Sorting
            sort_mapping = {
                "full_name": "user__full_name",
                "email": "user__email",
                "grade": "grade",
                "section": "section",
                "field": "field",
                "account_status": "user__is_active",
                "progress_percentage": "progress_percentage",
                "total_courses": "total_courses",
            }
            sort_field = sort_by.lstrip("-")
            db_sort_field = sort_mapping.get(sort_field, "user__date_joined")
            profiles = profiles.order_by(
                f"-{db_sort_field}" if sort_by.startswith("-") else db_sort_field
            )

            # Pagination
            total_count = profiles.count()
            total_pages = math.ceil(total_count / page_size) if page_size > 0 else 1
            start_index = (page - 1) * page_size
            end_index = start_index + page_size
            paginated_profiles = profiles[start_index:end_index]

            # Helper for progress rating
            def get_progress_rating(percentage):
                if percentage >= 90:
                    return "excellent"
                if percentage >= 70:
                    return "good"
                if percentage >= 40:
                    return "average"
                return "poor"

            # Build student data
            students_data = []
            for profile in paginated_profiles:
                user = profile.user
                profile_pic_url, _ = cloudinary.utils.cloudinary_url(
                    user.profile_pic_id,
                    resource_type="image",
                    type="authenticated",
                    sign_url=True,
                    secure=True,
                )

                progress_percentage = round(profile.progress_percentage or 0, 2)
                progress_points = (profile.finished_courses or 0) * 10
                progress_rating = get_progress_rating(progress_percentage)

                students_data.append(
                    {
                        "id": user.id,
                        "full_name": user.full_name or "",
                        "email": user.email or "",
                        "grade": profile.grade or "",
                        "section": profile.section or "",
                        "field": profile.field or "",
                        "profile_pic_url": profile_pic_url or "",
                        "account_status": "active" if user.is_active else "inactive",
                        "progress": {
                            "total_courses": profile.total_courses or 0,
                            "started_courses": profile.started_courses or 0,
                            "finished_courses": profile.finished_courses or 0,
                            "progress_percentage": progress_percentage,
                            "progress_points": progress_points,
                            "progress_rating": progress_rating,
                        },
                    }
                )

            # Global stats
            total_finished = UserCourseProgress.objects.filter(
                status="finished"
            ).count()
            total_progress_records = UserCourseProgress.objects.count()
            average_progress = (
                (total_finished / total_progress_records * 100)
                if total_progress_records > 0
                else 0
            )

            active_students = Profile.objects.filter(
                user__is_active=True, user__role="user"
            ).count()
            inactive_students = Profile.objects.filter(
                user__is_active=False, user__role="user"
            ).count()
            total_students = User.objects.filter(role="user").count()

            return Response(
                {
                    "students": students_data,
                    "pagination": {
                        "current_page": page,
                        "page_size": page_size,
                        "total_count": total_count,
                        "total_pages": total_pages,
                    },
                    "stats": {
                        "progress_avg": round(average_progress, 2),
                        "total": total_students,
                        "active": active_students,
                        "inactive": inactive_students,
                    },
                    "filter_options": filter_options,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            import traceback

            print(f"Error in StudentsView: {e}")
            print(traceback.format_exc())
            return Response(
                {"error": str(e), "detail": "Failed to fetch students"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@method_decorator(csrf_protect, name="dispatch")
class StudentUpdateView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAdminUser]

    def put(self, request, pk):
        try:
            user = User.objects.get(pk=pk, role="user")
            profile = Profile.objects.get(user=user)

            errors = {}

            email = (request.data.get("email") or "").strip()
            full_name = (request.data.get("full_name") or "").strip()
            gender = (request.data.get("gender") or "").strip().lower()
            grade = request.data.get("grade")
            section = (request.data.get("section") or "").strip()
            field = (request.data.get("field") or "").strip()
            account = (request.data.get("account") or "N/A").strip()
            phone_number = (request.data.get("phone_number") or "").strip()
            account_status = request.data.get("account_status", "active")
            profile_pic = request.FILES.get("profile_pic")

            if profile_pic:
                if profile_pic.size > 10 * 1024 * 1024:

                    errors["profile_pic"] = ["Profile picture must be less than 10MB"]

                if not profile_pic.content_type.startswith("image/"):
                    errors["profile_pic"] = ["Only image files are allowed"]

            if not email:
                errors["email"] = ["Email is required"]
            else:
                try:
                    validate_email(email)
                except ValidationError:
                    errors["email"] = ["Invalid email format"]

            if (
                email
                and user.email != email
                and User.objects.filter(email=email).exists()
            ):
                errors["email"] = ["User with this email already exists"]

            if not full_name:
                errors["full_name"] = ["Full name is required"]

            if gender not in ["male", "female", "other"]:
                errors["gender"] = ["Gender must be male, female, or other"]

            try:
                grade = int(grade)
                if grade < 1 or grade > 12:
                    raise ValueError
            except Exception:
                errors["grade"] = ["Grade must be between 1 and 12"]

            if not section or len(section) != 1 or not section.isalpha():
                errors["section"] = ["Section must be a single letter (A-Z)"]

            if not field:
                errors["field"] = ["Field is required"]

            if not phone_number:
                errors["phone_number"] = ["Phone number is required"]

            if errors:
                return Response(
                    {"detail": "Validation failed", "errors": errors},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # -------- atomic update --------
            with transaction.atomic():

                # update user (incl. profile_pic)
                user_serializer = UserSerializer(
                    user,
                    data={
                        "email": email,
                        "full_name": full_name,
                        "gender": gender,
                        "is_active": account_status == "active",
                    },
                    partial=True,
                    context={"request": request},
                )
                user_serializer.is_valid(raise_exception=True)
                user_serializer.save()

                # update profile
                profile.grade = grade
                profile.section = section.upper()
                profile.field = field
                profile.account = account
                profile.phone_number = phone_number
                profile.save()

            return Response(
                {
                    "message": "Student updated successfully",
                    "user": user_serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except User.DoesNotExist:
            return Response(
                {"detail": "Student not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Profile.DoesNotExist:
            return Response(
                {"detail": "Student profile not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentDetailView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAdminUser]

    def get(self, request, student_id):
        try:
            profile = Profile.objects.select_related("user").get(
                user__id=student_id, user__role="user"
            )
            user = profile.user

            profile_data = ProfileSerializer(profile).data

            # Profile picture
            profile_pic_url = None
            if user.profile_pic_id:
                try:
                    profile_pic_url, _ = cloudinary.utils.cloudinary_url(
                        user.profile_pic_id,
                        resource_type="image",
                        type="authenticated",
                        sign_url=True,
                        secure=True,
                    )
                except Exception:
                    profile_pic_url = None

            # Course progress annotations
            progress_stats = UserCourseProgress.objects.filter(user=user).aggregate(
                total_courses=Count("course", distinct=True),
                started_courses=Count("id", filter=Q(status="started")),
                finished_courses=Count("id", filter=Q(status="finished")),
            )

            total_courses = progress_stats.get("total_courses") or 0
            started_courses = progress_stats.get("started_courses") or 0
            finished_courses = progress_stats.get("finished_courses") or 0

            # Calculate progress percentage, points, rating
            progress_percentage = round(
                (finished_courses / total_courses * 100) if total_courses > 0 else 0.0,
                2,
            )
            progress_points = finished_courses * 10  # 10 points per finished course
            if progress_percentage >= 90:
                progress_rating = "excellent"
            elif progress_percentage >= 70:
                progress_rating = "good"
            elif progress_percentage >= 40:
                progress_rating = "average"
            else:
                progress_rating = "poor"

            student_data = {
                "profile": profile_data,
                "profile_pic_url": profile_pic_url,
                "progress_summary": {
                    "total_courses": total_courses,
                    "started_courses": started_courses,
                    "finished_courses": finished_courses,
                    "progress_percentage": progress_percentage,
                    "progress_points": progress_points,
                    "progress_rating": progress_rating,
                },
            }

            return Response({"student": student_data}, status=status.HTTP_200_OK)

        except Profile.DoesNotExist:
            return Response(
                {"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e), "detail": "Failed to fetch student details"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentDataView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, student_id):
        try:
            student = User.objects.get(id=student_id)
            student_serializer = UserInverseSerializer(student)
            student_data = student_serializer.data

            # Calculate progress
            progress_stats = UserCourseProgress.objects.filter(user=student).aggregate(
                total_courses=Count("course", distinct=True),
                started_courses=Count("id", filter=Q(status="started")),
                finished_courses=Count("id", filter=Q(status="finished")),
            )

            total_courses = progress_stats.get("total_courses") or 0
            started_courses = progress_stats.get("started_courses") or 0
            finished_courses = progress_stats.get("finished_courses") or 0

            progress_percentage = round(
                (finished_courses / total_courses * 100) if total_courses > 0 else 0.0,
                2,
            )
            progress_points = finished_courses * 10
            if progress_percentage >= 90:
                progress_rating = "excellent"
            elif progress_percentage >= 70:
                progress_rating = "good"
            elif progress_percentage >= 40:
                progress_rating = "average"
            else:
                progress_rating = "poor"

            return Response(
                {
                    "student": {
                        "id": student_data["id"],
                        "full_name": student_data["full_name"],
                        "email": student_data["email"],
                        "gender": student_data.get("gender"),
                        "grade": student_data["profile"].get("grade"),
                        "section": student_data["profile"].get("section"),
                        "field": student_data["profile"].get("field"),
                        "phone_number": student_data["profile"].get("phone_number"),
                        "account": student_data["profile"].get("account"),
                        "account_status": (
                            "active" if student_data["is_active"] else "inactive"
                        ),
                        "profile_pic_url": student_data.get("profile_pic_url"),
                        "progress_summary": {
                            "total_courses": total_courses,
                            "started_courses": started_courses,
                            "finished_courses": finished_courses,
                            "progress_percentage": progress_percentage,
                            "progress_points": progress_points,
                            "progress_rating": progress_rating,
                        },
                    }
                },
                status=status.HTTP_200_OK,
            )

        except User.DoesNotExist:
            return Response(
                {"error": "User not found."}, status=status.HTTP_404_NOT_FOUND
            )


@method_decorator(csrf_protect, name="dispatch")
class StudentDeleteView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_object(self, pk):
        try:
            return User.objects.get(pk=pk, role="user")
        except User.DoesNotExist:
            return None

    def delete(self, request, pk):
        try:
            user = self.get_object(pk)
            if not user:
                return Response(
                    {"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND
                )

            student_data = {
                "id": user.id,
                "full_name": user.full_name,
                "email": user.email,
            }

            with transaction.atomic():
                Profile.objects.filter(user=user).delete()
                user.delete()

            return Response(
                {
                    "message": "Student deleted successfully",
                    "deleted_student": student_data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            import traceback

            print(f"Error deleting student: {str(e)}")
            print(traceback.format_exc())

            return Response(
                {
                    "error": "Failed to delete student",
                    "detail": "An unexpected error occurred",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@method_decorator(csrf_protect, name="dispatch")
class StudentCreateView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUser]
    FIELD_LIST = ["ai", "other", "backend", "frontend", "embedded", "cyber"]

    def post(self, request):
        try:
            email = (request.data.get("email") or "").strip()
            full_name = (request.data.get("full_name") or "").strip()
            grade = request.data.get("grade")
            section = (request.data.get("section") or "").strip().upper()
            field = (request.data.get("field") or "").strip().lower()
            account = (request.data.get("account") or "N/A").strip()
            phone_number = (request.data.get("phone_number") or "").strip()

            # Validate all required fields
            errors = {}

            if not email:
                errors["email"] = ["Email is required"]
            else:
                try:
                    validate_email(email)
                except ValidationError:
                    errors["email"] = ["Invalid email format"]

            if not full_name:
                errors["full_name"] = ["Full name is required"]

            if grade is None:
                errors["grade"] = ["Grade is required"]
            else:
                try:
                    grade = int(grade)
                    if grade < 1 or grade > 12:
                        errors["grade"] = ["Grade must be between 1 and 12"]
                except (ValueError, TypeError):
                    errors["grade"] = ["Grade must be a valid number"]

            if not section:
                errors["section"] = ["Section is required"]
            elif len(section) != 1 or not section.isalpha():
                errors["section"] = ["Section must be a single letter (A-Z)"]

            if not field:
                errors["field"] = ["Field is required"]

            if field not in self.FIELD_LIST:
                errors["field"] = ["Invalid field name"]

            if not phone_number:
                errors["phone_number"] = ["Phone number is required"]

            # Check for existing email
            if (
                email
                and not errors.get("email")
                and User.objects.filter(email=email).exists()
            ):
                errors["email"] = ["User with this email already exists"]

            if errors:
                return Response(
                    {"detail": "Validation failed", "errors": errors},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            with transaction.atomic():
                # Create user
                user = User.objects.create_user(
                    email=email,
                    full_name=full_name,
                    is_active=True,  # Always active on creation
                    role="user",
                )

                # Create profile with proper field values
                profile = Profile.objects.create(
                    user=user,
                    grade=grade,
                    section=section,
                    field=field,
                    account=account,
                    phone_number=phone_number,
                )

            response_data = {
                "id": user.id,
                "full_name": user.full_name,
                "email": user.email,
                "grade": profile.grade,
                "section": profile.section,
                "field": profile.field,
                "account": profile.account,
                "phone_number": profile.phone_number,
                "account_status": "active",
                "created_at": profile.created_at,
                "message": "Student created successfully",
            }

            return Response(response_data, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            return Response(
                {"detail": "Validation error", "errors": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            import traceback

            print(f"Error creating student: {str(e)}")
            print(traceback.format_exc())

            return Response(
                {
                    "detail": "Failed to create student",
                    "error": "An unexpected error occurred",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@method_decorator(csrf_protect, name="dispatch")
class StudentsBulkUploadView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUser]
    FIELD_LIST = ["ai", "other", "backend", "frontend", "embedded", "cyber"]

    def post(self, request):
        try:
            if "file" not in request.FILES:
                return Response(
                    {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
                )

            file = request.FILES["file"]
            file_extension = file.name.split(".")[-1].lower()

            # ---------- READ FILE WITHOUT PANDAS ----------
            try:
                if file_extension == "csv":
                    decoded = file.read().decode("utf-8")
                    reader = csv.DictReader(io.StringIO(decoded))
                    rows = list(reader)
                    df = SimpleDataFrame(rows, reader.fieldnames)

                elif file_extension in ["xlsx", "xls"]:
                    wb = load_workbook(file)
                    sheet = wb.active

                    headers = [cell.value for cell in sheet[1]]
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row))
                        rows.append(row_dict)

                    df = SimpleDataFrame(rows, headers)

                else:
                    return Response(
                        {"error": "Unsupported file format. Use CSV or Excel"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            except Exception as e:
                return Response(
                    {"error": f"Failed to read file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # ---------- VALIDATE REQUIRED COLUMNS ----------
            required_columns = [
                "full_name",
                "email",
                "grade",
                "section",
                "field",
                "gender",
                "phone_number",
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                return Response(
                    {
                        "error": f"Missing required columns: {', '.join(missing_columns)}",
                        "required_columns": required_columns,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            created_students = []
            errors = []
            users_to_create = []
            profiles_to_create = []

            emails_in_file = set()
            duplicate_emails_in_file = set()

            # ---------- FIRST PASS: VALIDATION ----------
            for index, row in df.iterrows():
                row_errors = {}

                email = str(row["email"]).strip().lower()
                full_name = str(row.get("full_name", "")).strip()
                grade_str = str(row.get("grade", "")).strip()
                section = str(row.get("section", "")).strip().upper()
                field = str(row.get("field", "")).strip().lower()
                phone_number = str(row.get("phone_number", "")).strip()
                account = str(row.get("account", "")).strip()
                gender = str(row.get("gender", "male")).strip().lower()

                if gender not in ["male", "female"]:
                    row_errors["gender"] = ["Invalid gender"]

                if email in emails_in_file:
                    duplicate_emails_in_file.add(email)
                emails_in_file.add(email)

                if not email:
                    row_errors["email"] = ["Email is required"]
                else:
                    try:
                        validate_email(email)
                    except ValidationError:
                        row_errors["email"] = ["Invalid email format"]

                if not full_name:
                    row_errors["full_name"] = ["Full name is required"]

                if not grade_str:
                    row_errors["grade"] = ["Grade is required"]
                else:
                    try:
                        grade = int(grade_str)
                        if grade < 1 or grade > 12:
                            row_errors["grade"] = ["Grade must be between 1 and 12"]
                    except (ValueError, TypeError):
                        row_errors["grade"] = ["Grade must be a valid number"]

                if not section:
                    row_errors["section"] = ["Section is required"]
                elif len(section) != 1 or not section.isalpha():
                    row_errors["section"] = ["Section must be a single letter (A-Z)"]

                if not field:
                    row_errors["field"] = ["Field is required"]
                elif field not in self.FIELD_LIST:
                    row_errors["field"] = [f"'{field}' is invalid field name."]

                if not phone_number:
                    row_errors["phone_number"] = ["Phone number is required"]

                if row_errors:
                    errors.append(
                        {"row": index + 1, "email": email, "errors": row_errors}
                    )

            if duplicate_emails_in_file:
                for index, row in df.iterrows():
                    email = str(row["email"]).strip().lower()
                    if email in duplicate_emails_in_file:
                        errors.append(
                            {
                                "row": index + 1,
                                "email": email,
                                "errors": {"email": ["Duplicate email in the file"]},
                            }
                        )

            if errors:
                return Response(
                    {
                        "error": "Validation failed for some rows",
                        "created_count": 0,
                        "created_students": [],
                        "error_count": len(errors),
                        "errors": errors,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # ---------- SECOND PASS: CREATE USERS ----------
            with transaction.atomic():
                existing_emails = set(
                    User.objects.filter(
                        email__in=[
                            str(row["email"]).strip().lower()
                            for _, row in df.iterrows()
                        ]
                    ).values_list("email", flat=True)
                )

                for index, row in df.iterrows():
                    email = str(row["email"]).strip().lower()
                    full_name = str(row.get("full_name", "")).strip()
                    gender = str(row.get("gender", "male")).strip().lower()

                    if email in existing_emails:
                        errors.append(
                            {
                                "row": index + 1,
                                "email": email,
                                "error": "User with this email already exists in the database",
                            }
                        )
                        continue

                    user = User(
                        email=email,
                        full_name=full_name,
                        is_active=True,
                        role="user",
                        gender=gender,  # ← gender now goes to User
                    )
                    users_to_create.append(user)

                if users_to_create:
                    created_users = User.objects.bulk_create(users_to_create)

                    user_dict = {user.email: user for user in created_users}

                    for index, row in df.iterrows():
                        email = str(row["email"]).strip().lower()
                        if email in user_dict:
                            user = user_dict[email]
                            grade_str = str(row.get("grade", "")).strip()
                            section = str(row.get("section", "")).strip().upper()
                            field = str(row.get("field", "")).strip()
                            phone_number = str(row.get("phone_number", "")).strip()
                            account = str(row.get("account", "")).strip()

                            profile = Profile(
                                user=user,
                                grade=int(grade_str),
                                section=section.upper(),
                                field=field.lower(),
                                account=(
                                    account if account and account != "N/A" else None
                                ),
                                phone_number=phone_number,
                            )
                            profiles_to_create.append(profile)

                    if profiles_to_create:
                        Profile.objects.bulk_create(profiles_to_create)

                    for user in created_users:
                        profile = next(
                            (p for p in profiles_to_create if p.user_id == user.id),
                            None,
                        )
                        created_students.append(
                            {
                                "id": user.id,
                                "full_name": user.full_name,
                                "email": user.email,
                                "grade": profile.grade if profile else None,
                                "section": profile.section if profile else None,
                                "field": profile.field if profile else None,
                                "account": profile.account if profile else "N/A",
                                "phone_number": (
                                    profile.phone_number if profile else None
                                ),
                                "gender": user.gender,  # ← from User
                                "message": "Please change your password on first login",
                            }
                        )

            response_data = {
                "created_count": len(created_students),
                "created_students": created_students,
                "error_count": len(errors),
                "errors": errors,
            }

            status_code = (
                status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED
            )
            return Response(response_data, status=status_code)

        except Exception as e:
            import traceback

            traceback.print_exc()
            return Response(
                {"error": str(e), "detail": "Failed to process bulk upload"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentsExportView(APIView):

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        try:
            # Get query parameters for filtering
            search = request.query_params.get("search", "").strip()
            grade = request.query_params.get("grade", "").strip()
            section = request.query_params.get("section", "").strip()
            account_status = request.query_params.get("account_status", "").strip()

            # Get filtered students
            profiles = Profile.objects.select_related("user").filter(
                user__role="user", user__is_deleted=False
            )

            if search:
                profiles = profiles.filter(
                    Q(user__full_name__icontains=search)
                    | Q(user__email__icontains=search)
                    | Q(grade__icontains=search)
                    | Q(section__icontains=search)
                )

            if grade:
                try:
                    grade_int = int(grade)
                    profiles = profiles.filter(grade=grade_int)
                except ValueError:
                    pass

            if section:
                profiles = profiles.filter(section=section)

            if account_status:
                if account_status == "active":
                    profiles = profiles.filter(user__is_active=True)
                elif account_status == "inactive":
                    profiles = profiles.filter(user__is_active=False)

            # ---------- PREPARE EXCEL ----------
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"

            # Column headers
            headers = [
                "Full Name",
                "Email",
                "Grade",
                "Section",
                "Field",
                "Account",
                "Phone Number",
                "Gender",
                "Account Status",
                "Created At",
                "Last Login",
                "Date Joined",
            ]
            ws.append(headers)

            # Data rows
            for profile in profiles:
                ws.append(
                    [
                        profile.user.full_name,
                        profile.user.email,
                        profile.grade or "",
                        profile.section or "",
                        profile.field or "",
                        profile.account or "",
                        profile.phone_number or "",
                        profile.user.gender or "",
                        "Active" if profile.user.is_active else "Inactive",
                        profile.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                        (
                            profile.user.last_login.strftime("%Y-%m-%d %H:%M:%S")
                            if profile.user.last_login
                            else ""
                        ),
                        profile.user.date_joined.strftime("%Y-%m-%d %H:%M:%S"),
                    ]
                )

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"students_export_{timestamp}.xlsx"

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {"error": str(e), "detail": "Failed to export students"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentsStatsView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        try:
            # Basic stats
            total_students = Profile.objects.count()
            active_students = Profile.objects.filter(user__is_active=True).count()
            inactive_students = Profile.objects.filter(user__is_active=False).count()

            # Grade distribution
            grade_distribution = {}
            grade_counts = (
                Profile.objects.values("grade")
                .annotate(count=Count("grade"))
                .order_by("grade")
            )
            for item in grade_counts:
                if item["grade"]:
                    grade_distribution[str(item["grade"])] = item["count"]

            # Section distribution
            section_distribution = {}
            section_counts = (
                Profile.objects.values("section")
                .annotate(count=Count("section"))
                .order_by("section")
            )
            for item in section_counts:
                if item["section"]:
                    section_distribution[item["section"]] = item["count"]

            # Field distribution
            field_distribution = {}
            field_counts = (
                Profile.objects.values("field")
                .annotate(count=Count("field"))
                .order_by("field")
            )
            for item in field_counts:
                if item["field"]:
                    field_distribution[item["field"]] = item["count"]

            # Account distribution
            account_distribution = {}
            account_counts = (
                Profile.objects.values("account")
                .annotate(count=Count("account"))
                .order_by("account")
            )
            for item in account_counts:
                if item["account"]:
                    account_distribution[item["account"]] = item["count"]

            # Monthly registration trend (last 12 months)
            twelve_months_ago = timezone.now() - timezone.timedelta(days=365)
            monthly_trend = (
                Profile.objects.filter(created_at__gte=twelve_months_ago)
                .annotate(month=TruncMonth("created_at"))
                .values("month")
                .annotate(count=Count("id"))
                .order_by("month")
            )

            trend_data = [
                {"month": item["month"].strftime("%Y-%m"), "count": item["count"]}
                for item in monthly_trend
            ]

            # Progress stats
            total_progress_records = UserCourseProgress.objects.count()
            total_finished_courses = UserCourseProgress.objects.filter(
                status="finished"
            ).count()
            average_progress_percentage = (
                round((total_finished_courses / total_progress_records * 100), 2)
                if total_progress_records > 0
                else 0
            )

            # Optional: distribution by field (average progress per field)
            field_progress_distribution = {}
            for field in (
                Profile.objects.exclude(field__isnull=True)
                .exclude(field="")
                .values_list("field", flat=True)
                .distinct()
            ):
                field_students = Profile.objects.filter(field=field)
                user_ids = field_students.values_list("user__id", flat=True)
                finished_count = UserCourseProgress.objects.filter(
                    user__id__in=user_ids, status="finished"
                ).count()
                total_count = UserCourseProgress.objects.filter(
                    user__id__in=user_ids
                ).count()
                avg_progress = (
                    round((finished_count / total_count * 100), 2)
                    if total_count > 0
                    else 0
                )
                field_progress_distribution[field] = avg_progress

            response_data = {
                "overall": {
                    "total_students": total_students,
                    "active_students": active_students,
                    "inactive_students": inactive_students,
                    "active_percentage": (
                        round((active_students / total_students * 100), 2)
                        if total_students > 0
                        else 0
                    ),
                    "average_progress_percentage": average_progress_percentage,
                },
                "distributions": {
                    "by_grade": grade_distribution,
                    "by_section": section_distribution,
                    "by_field": field_distribution,
                    "by_account": account_distribution,
                    "average_progress_by_field": field_progress_distribution,
                },
                "trends": {"monthly_registrations": trend_data},
                "last_updated": timezone.now().isoformat(),
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": str(e), "detail": "Failed to fetch statistics"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentTemplateView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        try:
            # Column names
            columns = [
                "full_name",
                "email",
                "grade",
                "section",
                "field",
                "phone_number",
                "account",
                "gender",
            ]

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Student Template"

            # Append column headers only
            ws.append(columns)

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                'attachment; filename="student_template.xlsx"'
            )

            return response

        except Exception as e:
            return Response(
                {"error": str(e), "detail": "Failed to download template"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GetAllUsersView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAdminUser]

    def get(self, request):
        users = Profile.objects.filter(user__role="user")
        if not users.exists():
            return Response(
                {"warning": "No user found."}, status=status.HTTP_400_BAD_REQUEST
            )

        serializer = ProfileSerializer(users, many=True)
        return Response({"users": serializer.data}, status=status.HTTP_200_OK)


class DashboardView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAdminUser]

    def get(self, request):
        students = User.objects.filter(role="user", is_deleted=False)

        # Gender distribution
        gender_counts_query = students.values("gender").annotate(count=Count("id"))
        gender_counts = {item["gender"]: item["count"] for item in gender_counts_query}
        for g in ["male", "female"]:
            gender_counts.setdefault(g, 0)

        # Total students
        total_students = students.count()

        # Progress stats
        total_progress_records = UserCourseProgress.objects.count()
        total_finished_courses = UserCourseProgress.objects.filter(
            status="finished"
        ).count()
        average_progress_percentage = round(
            (
                (total_finished_courses / total_progress_records * 100)
                if total_progress_records > 0
                else 0.0
            ),
            2,
        )

        # Grade distribution
        grade_distribution_query = (
            Profile.objects.exclude(user__role="admin")
            .values("grade")
            .annotate(count=Count("id"))
            .order_by("grade")
        )
        grade_distribution = {
            item["grade"]: item["count"]
            for item in grade_distribution_query
            if item["grade"]
        }

        # Average progress per grade
        progress_by_grade = {}
        for item in grade_distribution_query:
            grade = item["grade"]
            if not grade:
                continue
            grade_students = Profile.objects.filter(grade=grade).values_list(
                "user__id", flat=True
            )
            total_courses = UserCourseProgress.objects.filter(
                user__id__in=grade_students
            ).count()
            finished_courses = UserCourseProgress.objects.filter(
                user__id__in=grade_students, status="finished"
            ).count()
            avg_progress = (
                round((finished_courses / total_courses * 100), 2)
                if total_courses > 0
                else 0
            )
            progress_by_grade[grade] = avg_progress

        response_data = {
            "gender_counts": gender_counts,
            "total_students": total_students,
            "average_progress_percentage": average_progress_percentage,
            "grade_distribution": grade_distribution,
            "progress_by_grade": progress_by_grade,
        }

        return Response(response_data)


@method_decorator(csrf_protect, name="dispatch")
class AdminControlView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsSuperUser]

    def get(self, request, pk=None):
        if pk:
            admin = (
                User.objects.filter(role="admin", is_staff=True, pk=pk)
                .exclude(id=request.user.id)
                .first()
            )
            if not admin:
                return Response(
                    {"error": "Admin not found."}, status=status.HTTP_404_NOT_FOUND
                )
            serializer = UserInverseSerializer(admin)
            return Response({"admin": serializer.data}, status=status.HTTP_200_OK)

        admins = User.objects.filter(role="admin", is_staff=True).exclude(
            id=request.user.id
        )
        admins_serializer = UserInverseSerializer(admins, many=True)
        return Response({"admins": admins_serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            email = (request.data.get("email") or "").strip()
            full_name = (request.data.get("full_name") or "").strip()
            grade = request.data.get("grade")
            section = (request.data.get("section") or "").strip()
            field = (request.data.get("field") or "").strip()
            account = (request.data.get("account") or "N/A").strip()
            phone_number = (request.data.get("phone_number") or "").strip()
            is_superuser = request.data.get("is_superuser", False)

            # Validate all required fields
            errors = {}

            if not email:
                errors["email"] = ["Email is required"]
            else:
                try:
                    validate_email(email)
                except ValidationError:
                    errors["email"] = ["Invalid email format"]

            if not full_name:
                errors["full_name"] = ["Full name is required"]

            if grade is None:
                errors["grade"] = ["Grade is required"]
            else:
                try:
                    grade = int(grade)
                    if grade < 1 or grade > 12:
                        errors["grade"] = ["Grade must be between 1 and 12"]
                except (ValueError, TypeError):
                    errors["grade"] = ["Grade must be a valid number"]

            if not section:
                errors["section"] = ["Section is required"]
            elif len(section) != 1 or not section.isalpha():
                errors["section"] = ["Section must be a single letter (A-Z)"]

            if not field:
                errors["field"] = ["Field is required"]

            if not phone_number:
                errors["phone_number"] = ["Phone number is required"]

            # Check for existing email
            if (
                email
                and not errors.get("email")
                and User.objects.filter(email=email).exists()
            ):
                errors["email"] = ["User with this email already exists"]

            if errors:
                return Response(
                    {"detail": "Validation failed", "errors": errors},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            with transaction.atomic():
                # Create user
                user = User.objects.create_user(
                    email=email,
                    full_name=full_name,
                    is_active=True,
                    is_staff=True,
                    is_superuser=is_superuser,
                    role="admin",
                )

                # Create profile with proper field values
                profile = Profile.objects.create(
                    user=user,
                    grade=grade,
                    section=section.upper() if section else None,
                    field=field if field else None,
                    account=account,
                    phone_number=phone_number if phone_number else None,
                )

            response_data = {
                "id": user.id,
                "full_name": user.full_name,
                "email": user.email,
                "grade": profile.grade,
                "section": profile.section,
                "field": profile.field,
                "account": profile.account,
                "phone_number": profile.phone_number,
                "account_status": "active",
                "created_at": profile.created_at,
                "message": "Admin created successfully",
            }

            return Response(response_data, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            return Response(
                {"detail": "Validation error", "errors": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            # Log the error for debugging
            import traceback

            print(f"Error creating admin: {str(e)}")
            print(traceback.format_exc())

            return Response(
                {
                    "detail": "Failed to create admin",
                    "error": "An unexpected error occurred",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def put(self, request, pk):
        try:
            admin = User.objects.get(pk=pk, role="admin")

            old_email = admin.email
            profile = Profile.objects.get(user=admin)

            errors = {}

            email = (request.data.get("email") or "").strip()
            full_name = (request.data.get("full_name") or "").strip()
            gender = (request.data.get("gender") or "").strip().lower()
            grade = request.data.get("grade")
            section = (request.data.get("section") or "").strip()
            field = (request.data.get("field") or "").strip()
            account = (request.data.get("account") or "N/A").strip()
            phone_number = (request.data.get("phone_number") or "").strip()
            account_status = request.data.get("account_status", "active")
            profile_pic = request.FILES.get("profile_pic")

            if profile_pic:
                if profile_pic.size > 10 * 1024 * 1024:

                    errors["profile_pic"] = ["Profile picture must be less than 10MB"]

                if not profile_pic.content_type.startswith("image/"):
                    errors["profile_pic"] = ["Only image files are allowed"]

            if not email:
                errors["email"] = ["Email is required"]
            else:
                try:
                    validate_email(email)
                except ValidationError:
                    errors["email"] = ["Invalid email format"]

            if (
                email
                and admin.email != email
                and User.objects.filter(email=email).exists()
            ):
                errors["email"] = ["Admin with this email already exists"]

            if not full_name:
                errors["full_name"] = ["Full name is required"]

            if gender not in ["male", "female", "other"]:
                errors["gender"] = ["Gender must be male, female, or other"]

            try:
                grade = int(grade)
                if grade < 1 or grade > 12:
                    raise ValueError
            except Exception:
                errors["grade"] = ["Grade must be between 1 and 12"]

            if not section or len(section) != 1 or not section.isalpha():
                errors["section"] = ["Section must be a single letter (A-Z)"]

            if not field:
                errors["field"] = ["Field is required"]

            if not phone_number:
                errors["phone_number"] = ["Phone number is required"]

            if errors:
                return Response(
                    {"detail": "Validation failed", "errors": errors},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            with transaction.atomic():

                user_serializer = UserSerializer(
                    admin,
                    data={
                        "full_name": full_name,
                        "gender": gender,
                        "is_active": account_status == "active",
                    },
                    partial=True,
                    context={"request": request},
                )
                user_serializer.is_valid(raise_exception=True)
                user = user_serializer.save()

                if old_email != email:
                    user.email_verified = False
                    user.email = email
                    user.save(update_fields=["email_verified", "email"])

                # update profile
                profile.grade = grade
                profile.section = section.upper()
                profile.field = field
                profile.account = account
                profile.phone_number = phone_number
                profile.save()

            return Response(
                {
                    "message": "Admin updated successfully",
                    "user": user_serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except User.DoesNotExist:
            return Response(
                {"detail": "Admin not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Profile.DoesNotExist:
            return Response(
                {"detail": "Admin profile not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def delete(self, request, pk):
        try:
            admin = (
                User.objects.filter(role="admin", is_staff=True, id=pk)
                .exclude(id=request.user.id)
                .first()
            )
        except User.DoesNotExist:
            return Response(
                {"error": "Admin not found."}, status=status.HTTP_404_NOT_FOUND
            )

        admin.delete()
        return Response(
            {"message": "Admin successfully deleted."}, status=status.HTTP_200_OK
        )


@method_decorator(csrf_protect, name="dispatch")
class SettingUpdateView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsSuperUser]

    def get(self, request):
        setting, _ = Setting.objects.get_or_create(id=1)
        serializer = SettingSerializer(setting)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request):
        setting, created = Setting.objects.get_or_create(id=1)

        serializer = SettingSerializer(setting, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GradesRankExportPdfView(APIView):

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUser]

    MAX_TEXT_LENGTH = 30

    def truncate(self, text):
        if not text:
            return ""
        return (
            text
            if len(text) <= self.MAX_TEXT_LENGTH
            else text[: self.MAX_TEXT_LENGTH - 3] + "..."
        )

    def get(self, request):
        try:
            # ----------------------------
            # Filters
            # ----------------------------
            search = request.query_params.get("search", "").strip()
            grade_q = request.query_params.get("grade", "").strip()
            section_q = request.query_params.get("section", "").strip()
            field_q = request.query_params.get("field", "").strip()
            account_status = request.query_params.get("account_status", "").strip()

            profiles = Profile.objects.select_related("user").filter(
                user__role="user",
                user__is_deleted=False,
            )

            if search:
                profiles = profiles.filter(
                    Q(user__full_name__icontains=search)
                    | Q(user__email__icontains=search)
                    | Q(grade__icontains=search)
                    | Q(section__icontains=search)
                    | Q(field__icontains=search)
                )

            if grade_q:
                try:
                    profiles = profiles.filter(grade=int(grade_q))
                except ValueError:
                    pass

            if section_q:
                profiles = profiles.filter(section=section_q)

            if field_q:
                profiles = profiles.filter(field=field_q)

            if account_status == "active":
                profiles = profiles.filter(user__is_active=True)
            elif account_status == "inactive":
                profiles = profiles.filter(user__is_active=False)

            # ----------------------------
            # Score Annotation
            # ----------------------------
            # Use course progress: started = 1 point, finished = 2 points
            annotated_qs = (
                profiles.annotate(
                    started_count=Coalesce(
                        Count(
                            "user__usercourseprogress",
                            filter=Q(user__usercourseprogress__status="started"),
                        ),
                        0,
                    ),
                    finished_count=Coalesce(
                        Count(
                            "user__usercourseprogress",
                            filter=Q(user__usercourseprogress__status="finished"),
                        ),
                        0,
                    ),
                )
                .annotate(score=F("started_count") + F("finished_count") * 2)
                .order_by("-score", "user__full_name")
                .values("user__full_name", "grade", "section", "field", "score")
            )

            rows = list(annotated_qs)

            # ----------------------------
            # Dense Ranking
            # ----------------------------
            result_rows = []
            prev_score = None
            rank = 0

            for idx, row in enumerate(rows, start=1):
                score = int(row["score"]) if row["score"] else 0
                if prev_score is None or score != prev_score:
                    rank += 1
                    prev_score = score
                result_rows.append(
                    [
                        str(idx),
                        self.truncate(row["user__full_name"]),
                        row["grade"] or "",
                        row["section"] or "",
                        row["field"] or "",
                        str(score),
                        str(rank),
                    ]
                )

            # ----------------------------
            # Build PDF
            # ----------------------------
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30,
            )

            elements = []
            styles = getSampleStyleSheet()

            # Title
            elements.append(
                Paragraph("<b>Student Course Progress Ranking</b>", styles["Title"])
            )
            elements.append(Spacer(1, 0.2 * inch))

            # Metadata
            metadata_lines = [
                f"Total Students: {len(result_rows)}",
                f"Exported At: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}",
            ]
            for line in metadata_lines:
                elements.append(Paragraph(line, styles["Normal"]))

            elements.append(Spacer(1, 0.4 * inch))

            # Table
            table_data = [
                ["#", "Full Name", "Grade", "Section", "Field", "Score", "Rank"]
            ] + result_rows
            wrapped_data = [
                [Paragraph(str(cell), styles["Normal"]) for cell in row]
                for row in table_data
            ]

            available_width = doc.width
            table_width = available_width * 0.90  # 90%
            col_widths = [
                table_width * 0.06,  # #
                table_width * 0.28,  # Full Name
                table_width * 0.12,  # Grade
                table_width * 0.12,  # Section
                table_width * 0.12,  # Field
                table_width * 0.15,  # Score
                table_width * 0.15,  # Rank
            ]

            table = Table(wrapped_data, repeatRows=1, colWidths=col_widths)
            table.hAlign = "CENTER"
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e0e0")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ALIGN", (0, 1), (0, -1), "CENTER"),
                        ("ALIGN", (5, 1), (6, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )

            elements.append(table)
            doc.build(elements)
            buffer.seek(0)

            filename = f"student_course_progress_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            response = HttpResponse(buffer.read(), content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {"error": "Failed to export ranking", "detail": str(e)},
                status=500,
            )


@method_decorator(csrf_protect, name="dispatch")
class StudentsBulkOperationView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsSuperUser]

    FIELD_LIST = ["ai", "other", "backend", "frontend", "embedded", "cyber"]

    def post(self, request):
        try:
            action = request.data.get("action")
            if action not in ["delete", "edit"]:
                return Response(
                    {"error": "Invalid action. Must be 'delete' or 'edit'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # --- Filters ---
            search = (request.data.get("search") or "").strip()
            grade_q = request.data.get("grade")
            section_q = (request.data.get("section") or "").strip().upper()
            field_q = (request.data.get("field") or "").strip().lower()
            account_status = (request.data.get("account_status") or "").strip().lower()

            profiles = Profile.objects.select_related("user").filter(
                user__role="user", user__is_deleted=False
            )

            if search:
                profiles = profiles.filter(
                    Q(user__full_name__icontains=search)
                    | Q(user__email__icontains=search)
                    | Q(grade__icontains=search)
                    | Q(section__icontains=search)
                    | Q(field__icontains=search)
                )

            if grade_q is not None:
                try:
                    grade_int = int(grade_q)
                    profiles = profiles.filter(grade=grade_int)
                except ValueError:
                    pass

            if section_q:
                profiles = profiles.filter(section__iexact=section_q)

            if field_q:
                profiles = profiles.filter(field__iexact=field_q)

            if account_status:
                if account_status == "active":
                    profiles = profiles.filter(user__is_active=True)
                elif account_status == "inactive":
                    profiles = profiles.filter(user__is_active=False)

            affected_count = profiles.count()
            if affected_count == 0:
                return Response(
                    {"message": "No students match the given filters."},
                    status=status.HTTP_200_OK,
                )

            with transaction.atomic():
                if action == "delete":
                    user_ids = list(
                        profiles.values_list("user_id", flat=True).distinct()
                    )
                    # Count users BEFORE deletion
                    user_deleted_count = len(user_ids)
                    # Delete users (cascades will still happen, but not counted)
                    deleted_detail = User.objects.filter(id__in=user_ids).delete()

                    return Response(
                        {
                            "deleted_count": user_deleted_count,
                            "cascaded_deletions": deleted_detail,  # optional debug info
                        },
                        status=status.HTTP_200_OK,
                    )

                elif action == "edit":
                    edit_data = request.data.get("edit_data", {})
                    allowed_fields = [
                        "grade",
                        "section",
                        "field",
                        "account",
                        "phone_number",
                        "is_active",
                    ]

                    update_data = {}
                    profile_update_data = {}

                    for key, value in edit_data.items():
                        if key not in allowed_fields:
                            continue
                        if key == "field" and value.lower() not in self.FIELD_LIST:
                            continue
                        if key == "section":
                            value = value.upper() if value else value
                        if key in [
                            "grade",
                            "section",
                            "field",
                            "account",
                            "phone_number",
                        ]:
                            profile_update_data[key] = value
                        if key == "is_active":
                            update_data["is_active"] = bool(value)

                    # Update users
                    if update_data:
                        user_ids = profiles.values_list("user_id", flat=True)
                        User.objects.filter(id__in=user_ids).update(**update_data)

                    # Update profiles
                    if profile_update_data:
                        profiles.update(**profile_update_data)

                    return Response(
                        {
                            "message": f"{affected_count} students updated successfully",
                            "updated_fields": list(profile_update_data.keys())
                            + list(update_data.keys()),
                        },
                        status=status.HTTP_200_OK,
                    )

        except Exception as e:
            import traceback

            traceback.print_exc()
            return Response(
                {"error": str(e), "detail": "Failed to perform bulk operation"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
